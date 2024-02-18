import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from PIL import Image
import io
import argparse

def add_file_to_spreadsheet(file_path, start_directory, worksheet, embed_images, row_height):
    relative_path = os.path.relpath(file_path, start_directory)
    clickable_url = f"file:///{os.path.abspath(os.path.dirname(file_path)).replace(os.sep, '/')}"
    filetype = os.path.splitext(file_path)[1][1:].upper() if os.path.splitext(file_path)[1] else 'Unknown'

    try:
        size = os.path.getsize(file_path)
    except FileNotFoundError:
        return

    row = worksheet.max_row + 1
    worksheet.cell(row=row, column=1, value=os.path.dirname(relative_path))
    worksheet.cell(row=row, column=2, value=os.path.basename(file_path))
    worksheet.cell(row=row, column=3, value=filetype)
    worksheet.cell(row=row, column=4, value=clickable_url).hyperlink = clickable_url
    worksheet.cell(row=row, column=4).style = "Hyperlink"
    worksheet.cell(row=row, column=5, value=size)

    print(f"Adding file to spreadsheet: {file_path}")  # Print file info being added to the spreadsheet

    print(f"Converting image: {os.path.basename(file_path)}")  # Print filename before image conversion
    
    if embed_images and filetype.lower() in ['png', 'jpg', 'jpeg', 'gif', 'bmp']:
        try:
            img = Image.open(file_path)
            # Convert palette mode (mode 'P') images to RGB mode
            if img.mode == 'P':
                print(f"Skipping image {os.path.basename(file_path)} with mode '{img.mode}' (Palette mode)")
                return  # Skip saving as JPEG
            # Convert RGBA mode images to RGB mode
            if img.mode == 'RGBA':
                print(f"Converting RGBA mode image to RGB mode: {file_path}")
                img = img.convert('RGB')
            max_row_height = row_height
            width, height = img.size
            scale_factor = max_row_height / height
            resized_img = img.resize((int(width * scale_factor), int(height * scale_factor)))
            img_bytes = io.BytesIO()
            img_format = filetype.lower()
            resized_img = img.resize((int(width * scale_factor), int(height * scale_factor)))
            img_bytes = io.BytesIO()
            if img_format:
                try:
                    resized_img.save(img_bytes, format=img_format.upper())
                except KeyError:
                    img_format = 'JPEG'
                    resized_img.save(img_bytes, format=img_format)
                img_bytes.seek(0)
                img = ExcelImage(img_bytes)
                img.anchor = f'F{row}'
                worksheet.add_image(img)
                worksheet.row_dimensions[row].height = max_row_height
            else:
                print(f"Unsupported file format: {os.path.basename(file_path)}")
        except Exception as e:
            print(f"Error occurred while processing {os.path.basename(file_path)}: {e}")

def scan_directory(directory, start_directory, worksheet, embed_images, row_height):
    for root, dirs, files in os.walk(directory):
        for file in files:
            file_path = os.path.join(root, file)
            add_file_to_spreadsheet(file_path, start_directory, worksheet, embed_images, row_height)

def main():
    parser = argparse.ArgumentParser(description='Directory Reference Generator')
    parser.add_argument('-d', '--directory', help='Directory to scan', required=True)
    parser.add_argument('-o', '--output', help='Output filename', required=True)
    parser.add_argument('-i', '--images', action='store_true', help='Embed images')
    parser.add_argument('-rh', '--row_height', type=int, default=300, help='Row height for embedded images (default: 300)')
    args = parser.parse_args()

    start_directory = args.directory
    output_filename = args.output
    if not output_filename.endswith('.xlsx'):
        output_filename += '.xlsx'

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = os.path.splitext(os.path.basename(output_filename))[0]
    scan_directory(start_directory, start_directory, worksheet, args.images, args.row_height)
    # Set font for hyperlinks
    worksheet.cell(row=1, column=4).font = Font(color="0000FF", underline="single")
    workbook.save(filename=output_filename)

    print("Directory reference created successfully!")

if __name__ == "__main__":
    main()
